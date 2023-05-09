import openpyxl

workbook = openpyxl.load_workbook('D:/PycharmProjects/testing/project/saucedemo_project/TestData/Testdata.xlsx')

#here creating a function to find the number od rows in a sheet
def fetch_number_of_rows(Sheetname):
    sh = workbook[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = workbook[Sheetname]
    cell = sh.cell(int(row, cell))
    return cell.value

